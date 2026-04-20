#!/usr/bin/env python3
"""成本数据库工具 V3.2 (Python + sqlite3)

用法：python cost_db.py <command> [args]

命令：
  insert --日期 ...    插入记录（自动创建成本项 + 自动校验）
  update <id> <f> <v>  更新记录
  delete <id>          删除记录（级联删除）
  query "<sql>"        SQL 查询
  dashboard            生成看板
  project "<name>"     项目成本报表
  stats                统计
  convert-tax <id>     含税/税前换算
  convert <id>         单位换算
  items [list|add]     成本项管理
  units [list|add|...] 单位管理
"""

import os
import sys
import re
import json
import sqlite3
import time
from datetime import datetime

# ── Paths ──

HERE = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.environ.get('COST_DB_PATH', os.path.join(HERE, '成本数据.db'))
DASHBOARD_PATH = os.path.join(HERE, '成本查询.md')
EXPORT_PATH = os.path.join(HERE, '成本数据_export.json')
PENDING_DIR = os.path.dirname(os.path.abspath(DB_PATH))

# ── DB Helpers ──

_unit_map = None


def open_db():
    """Open SQLite connection with WAL mode and dict row factory."""
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.row_factory = sqlite3.Row
    return conn


def run_query(conn, sql, params=()):
    cur = conn.execute(sql, params)
    cols = [d[0] for d in cur.description] if cur.description else []
    rows = [dict(zip(cols, row)) for row in cur.fetchall()]
    return rows


def run_exec(conn, sql, params=()):
    conn.execute(sql, params)


def fmt(n):
    """Format number with 2 decimal places and comma separator (zh-CN style)."""
    try:
        return f"{float(n):,.2f}"
    except (ValueError, TypeError):
        return str(n)


# ── Unit Normalization ──

def load_unit_map(conn):
    global _unit_map
    if _unit_map is not None:
        return _unit_map
    _unit_map = {}
    try:
        rows = run_query(conn, 'SELECT unit, standard_form FROM unit_standard')
        for r in rows:
            _unit_map[r['unit']] = r['standard_form']
    except Exception:
        pass
    return _unit_map


def normalize_unit(unit, conn):
    umap = load_unit_map(conn)
    # Strip "元/" prefix — default settlement is RMB, no need to repeat
    unit = re.sub(r'^元/', '', unit)
    if unit in umap:
        return umap[unit]
    # Handle compound units like "台·月" → try mapping the base part
    m = re.match(r'^(.+\/)(.+)$', unit)
    if m and m.group(2) in umap:
        return umap[m.group(2)]
    return unit


# ── Field Name Mapping ──

FIELD_MAP = {
    '日期': 'date', '大类': 'category', '名称': 'name', '规格': 'spec',
    '单价': 'price', '单位': 'unit', '地区': 'location', '项目': 'project_name',
    '询价方式': 'price_type', '报价人': 'source_person', '状态': 'status',
    '备注': 'remark', '录入设备': 'input_device', '原始文件': 'source_file',
    '计税方式': 'tax_method', '换算来源': 'conversion_source', '换算公式': 'conversion_formula',
}

ALLOWED_UPDATE_FIELDS = [
    'date', 'price', 'unit', 'tax_method', 'price_type', 'source', 'source_person',
    'location', 'project_name', 'spec', 'status', 'input_device', 'source_file', 'raw_text',
    'is_composite', 'conversion_source', 'conversion_formula', 'remark',
]

# Pre-built parameterized SQL for each allowed update field
FIELD_UPDATE_SQL = {f: f'UPDATE cost_price SET {f} = ? WHERE id = ?' for f in ALLOWED_UPDATE_FIELDS}


def map_field(f):
    return FIELD_MAP.get(f, f)


# ── JSON Export ──

def export_json(conn):
    rows = run_query(conn, """
        SELECT cp.id, cp.date AS '日期', ci.category AS '大类', ci.name AS '名称',
               cp.spec AS '规格', cp.price AS '单价', cp.unit AS '单位',
               cp.location AS '地区', cp.project_name AS '项目',
               cp.price_type AS '询价方式', cp.source_person AS '报价人',
               cp.status AS '状态', cp.remark AS '备注',
               cp.input_device AS '录入设备', cp.source_file AS '原始文件',
               cp.tax_method AS '计税方式', cp.is_composite,
               cp.conversion_source AS '换算来源', cp.conversion_formula AS '换算公式',
               cp.legacy_id
        FROM cost_price cp JOIN cost_item ci ON cp.item_id = ci.id
        ORDER BY cp.date DESC, cp.id DESC
    """)
    data = {
        'updated': datetime.now().strftime('%Y/%m/%d %H:%M:%S'),
        'total': len(rows),
        'records': rows,
    }
    with open(EXPORT_PATH, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


# ── Item Resolution ──

def find_or_create_item(conn, name, category, unit):
    normed_unit = normalize_unit(unit, conn)
    parts = normed_unit.split('/')
    base_unit = parts[1] if len(parts) > 1 else normed_unit

    # Try to find existing
    existing = run_query(conn,
        'SELECT id FROM cost_item WHERE name = ? AND category = ?',
        (name, category))
    if existing:
        return existing[0]['id']

    # Find unit_id
    unit_id = None
    unit_rows = run_query(conn,
        'SELECT id FROM unit_standard WHERE unit = ? AND standard_form = ?',
        (base_unit, base_unit))
    if unit_rows:
        unit_id = unit_rows[0]['id']

    conn.execute(
        'INSERT INTO cost_item (name, category, unit_id) VALUES (?, ?, ?)',
        (name, category, unit_id))
    new_rows = run_query(conn,
        'SELECT id FROM cost_item WHERE name = ? AND category = ?',
        (name, category))
    global _unit_map
    _unit_map = None  # reset cache
    return new_rows[0]['id']


# ── Record Lookup ──

def find_price_id(conn, id_str):
    # Try integer ID
    try:
        int_id = int(id_str)
        rows = run_query(conn, 'SELECT id FROM cost_price WHERE id = ?', (int_id,))
        if rows:
            return int_id
    except ValueError:
        pass
    # Try legacy_id
    rows = run_query(conn,
        "SELECT id FROM cost_price WHERE legacy_id = ?", (id_str,))
    if rows:
        return rows[0]['id']
    return None


# ── Price Category Helper ──

def get_price_category(method):
    market = ['电话询价', '微信询价', '现场询价', '询价单']
    if method == '信息价':
        return '信息价'
    if method == '合同价':
        return '合同价'
    if method == '定额':
        return '定额'
    if method in market:
        return '市场价'
    return '其他'


# ── Validation Engine ──

def validate_price(conn, name, unit, price, category=''):
    """Check price against validation_rule table. Returns list of warnings."""
    warnings = []
    # Try name + unit + category first (most specific)
    rules = run_query(conn,
        'SELECT low_price, high_price FROM validation_rule WHERE name = ? AND unit = ? AND category = ?',
        (name, unit, category))
    # Fallback to name + unit (for names that only exist in one category)
    if not rules:
        rules = run_query(conn,
            'SELECT low_price, high_price FROM validation_rule WHERE name = ? AND unit = ?',
            (name, unit))
    if rules:
        r = rules[0]
        if price < r['low_price'] or price > r['high_price']:
            warnings.append(
                f'价格超限: {price} 超出 [{r["low_price"]}, {r["high_price"]}] 区间')
    return warnings


def check_duplicate(conn, item_id, date, price, source_person):
    """Check for duplicate records. Returns list of warnings."""
    warnings = []
    dupes = run_query(conn,
        'SELECT id FROM cost_price WHERE item_id = ? AND date = ? AND price = ? AND source_person = ?',
        (item_id, date, price, source_person))
    if dupes:
        warnings.append(f'疑似重复: 已存在相同日期+单价+报价人的记录 #{dupes[0]["id"]}')
    return warnings


def check_trend(conn, name, unit, location, current_price):
    """Check if recent prices show sustained increase >20%. Returns warnings."""
    warnings = []
    recent = run_query(conn, """
        SELECT cp.price FROM cost_price cp
        JOIN cost_item ci ON cp.item_id = ci.id
        WHERE ci.name = ? AND cp.unit = ? AND (cp.location = ? OR ? = '')
        ORDER BY cp.date DESC LIMIT 3
    """, (name, unit, location, location))
    if len(recent) == 3:
        p1, p2, p3 = recent[2]['price'], recent[1]['price'], recent[0]['price']
        if p1 < p2 < p3:
            increase_pct = (p3 - p1) / p1 * 100
            if increase_pct > 20:
                warnings.append(f'价格趋势异常: 最近3次持续上涨 {increase_pct:.1f}%')
    return warnings


# ── Safe Formula Evaluation ──

def _safe_eval_formula(formula_text):
    """Evaluate a simple arithmetic formula with only * and / operators.

    Supported formats: "N", "N * M", "N * M / K", "N / M"
    where N, M, K are decimal numbers. No parentheses, no addition, no subtraction.
    """
    if not re.match(r'^[\d\s\.\*/]+$', formula_text):
        raise ValueError(f'公式包含非法字符: {formula_text}')
    tokens = formula_text.strip().split()
    if not tokens:
        raise ValueError('空公式')
    result = float(tokens[0])
    i = 1
    while i < len(tokens):
        op = tokens[i]
        if op not in ('*', '/'):
            raise ValueError(f'不支持的运算符: {op}')
        if i + 1 >= len(tokens):
            raise ValueError(f'公式不完整: {formula_text}')
        operand = float(tokens[i + 1])
        if op == '*':
            result *= operand
        else:
            if operand == 0:
                raise ValueError('除数为零')
            result /= operand
        i += 2
    return result


# ── Public API (importable by api_server.py) ──

def insert_record(params: dict) -> int:
    """Insert a price record. Returns the new ID. Raises on error."""
    conn = open_db()
    try:
        name = params.get('名称') or params.get('name')
        category = params.get('大类') or params.get('category')
        price = float(params.get('单价') or params.get('price') or 0)
        unit = params.get('单位') or params.get('unit')
        date = params.get('日期') or params.get('date')

        if not name or not category or not price or not unit or not date:
            raise ValueError('缺少必填字段（日期/大类/名称/单价/单位）')

        item_id = find_or_create_item(conn, name, category, unit)
        normed_unit = normalize_unit(unit, conn)

        # Validate breakdown sum
        labor = params.get('人工费')
        material = params.get('材料费')
        equipment = params.get('机械费')
        labor = float(labor) if labor else None
        material = float(material) if material else None
        equipment = float(equipment) if equipment else None

        status = params.get('状态') or params.get('status') or '待核实'
        remark = params.get('备注') or params.get('remark') or ''

        if labor is not None or material is not None or equipment is not None:
            s = (labor or 0) + (material or 0) + (equipment or 0)
            if price > 0 and abs(s - price) / price > 0.05:
                print(f'警告：工料机合计 {s} 与单价 {price} 偏差超过5%，已标记待核实',
                      file=sys.stderr)
                status = '待核实'
                remark += f' 工料机合计{s}与单价偏差{((s - price) / price * 100):.1f}%'

        # ── Auto Validation ──
        source_person = params.get('报价人') or params.get('source_person') or ''
        location = params.get('地区') or params.get('location') or ''

        validation_warnings = []
        validation_warnings.extend(validate_price(conn, name, normed_unit, price, category))
        validation_warnings.extend(check_duplicate(conn, item_id, date, price, source_person))
        validation_warnings.extend(check_trend(conn, name, normed_unit, location, price))

        if validation_warnings:
            status = '待核实'
            for w in validation_warnings:
                remark += f' [{w}]'
                print(f'校验警告: {w}', file=sys.stderr)

        # Check if name exists in validation rules
        rule_check = run_query(conn,
            'SELECT id FROM validation_rule WHERE name = ? LIMIT 1', (name,))
        if not rule_check:
            remark += ' [待补充词条]'

        is_composite = 1 if category == '综合' else 0

        conn.execute("""
            INSERT INTO cost_price
            (item_id, price, unit, date, tax_method, price_type, source_person,
             location, project_name, spec, status, input_device, source_file,
             is_composite, conversion_source, conversion_formula, remark)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            item_id, price, normed_unit, date,
            params.get('计税方式') or params.get('tax_method') or '不详',
            params.get('询价方式') or params.get('price_type') or '',
            source_person,
            location,
            params.get('项目') or params.get('project_name') or '',
            params.get('规格') or params.get('spec') or '',
            status,
            params.get('录入设备') or params.get('input_device') or '',
            params.get('原始文件') or params.get('source_file') or '',
            is_composite,
            params.get('换算来源') or params.get('conversion_source') or '',
            params.get('换算公式') or params.get('conversion_formula') or '',
            remark,
        ))

        new_id = run_query(conn, 'SELECT last_insert_rowid() AS id')[0]['id']

        # Create components
        if labor is not None:
            conn.execute(
                'INSERT INTO cost_component (price_id, component_type, price) VALUES (?, ?, ?)',
                (new_id, '人工', labor))
        if material is not None:
            conn.execute(
                'INSERT INTO cost_component (price_id, component_type, price) VALUES (?, ?, ?)',
                (new_id, '材料', material))
        if equipment is not None:
            conn.execute(
                'INSERT INTO cost_component (price_id, component_type, price) VALUES (?, ?, ?)',
                (new_id, '机械', equipment))

        if normed_unit != unit:
            print(f'单位已归一化：{unit} → {normed_unit}')
        print(f'已插入 #{new_id} ({name}, {price} {normed_unit})')
        export_json(conn)
        conn.commit()
        return new_id
    except UnicodeEncodeError:
        # print 编码失败不应导致回滚，数据已准备好
        conn.commit()
        return new_id
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def update_record(id_str: str, field: str, value: str) -> bool:
    """Update a record field. Returns True on success."""
    field = map_field(field)
    conn = open_db()
    try:
        price_id = find_price_id(conn, id_str)
        if not price_id:
            raise ValueError(f'未找到记录 {id_str}')

        # Component types
        comp_types = {'人工费': '人工', '材料费': '材料', '机械费': '机械'}
        if field in comp_types:
            c_type = comp_types[field]
            num_val = float(value)
            existing = run_query(conn,
                'SELECT id FROM cost_component WHERE price_id = ? AND component_type = ?',
                (price_id, c_type))
            if existing:
                conn.execute(
                    'UPDATE cost_component SET price = ? WHERE id = ?',
                    (num_val, existing[0]['id']))
            else:
                conn.execute(
                    'INSERT INTO cost_component (price_id, component_type, price) VALUES (?, ?, ?)',
                    (price_id, c_type, num_val))
            conn.commit()
            print(f'已更新 #{price_id} {field} = {value}')
            return True

        if field not in ALLOWED_UPDATE_FIELDS:
            raise ValueError(f'不允许更新字段 "{field}"')

        numeric_fields = {'price', 'is_composite'}
        val = float(value) if field in numeric_fields else value
        conn.execute(FIELD_UPDATE_SQL[field], (val, price_id))
        conn.commit()
        print(f'已更新 #{price_id}.{field} = {value}')
        return True
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def delete_record(id_str: str) -> bool:
    """Delete a record with cascade. Returns True on success."""
    conn = open_db()
    try:
        price_id = find_price_id(conn, id_str)
        if not price_id:
            raise ValueError(f'未找到记录 {id_str}')
        conn.execute('DELETE FROM cost_feature WHERE price_id = ?', (price_id,))
        conn.execute('DELETE FROM cost_component WHERE price_id = ?', (price_id,))
        conn.execute('DELETE FROM cost_price WHERE id = ?', (price_id,))
        conn.commit()
        print(f'已删除 #{price_id}')
        return True
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def query_prices(sql: str) -> list:
    """Execute arbitrary SQL and return results."""
    conn = open_db()
    try:
        rows = run_query(conn, sql)
        return rows
    finally:
        conn.close()


def search_prices(keyword: str) -> list:
    """Search prices by keyword across name, spec, project."""
    conn = open_db()
    try:
        like = f'%{keyword}%'
        rows = run_query(conn, """
            SELECT cp.id, ci.category AS 大类, ci.name AS 名称, cp.spec AS 规格,
                   cp.price AS 单价, cp.unit AS 单位, cp.location AS 地区,
                   cp.date AS 日期, cp.status AS 状态, cp.project_name AS 项目,
                   cp.price_type AS 询价方式, cp.source_person AS 报价人
            FROM cost_price cp
            JOIN cost_item ci ON cp.item_id = ci.id
            WHERE ci.name LIKE ? OR cp.spec LIKE ? OR cp.project_name LIKE ?
            ORDER BY cp.date DESC
            LIMIT 20
        """, (like, like, like))
        return rows
    finally:
        conn.close()


def confirm_record(id_str: str) -> bool:
    """Change status from 待核实 to 已确认."""
    conn = open_db()
    try:
        price_id = find_price_id(conn, id_str)
        if not price_id:
            raise ValueError(f'未找到记录 {id_str}')
        conn.execute('UPDATE cost_price SET status = ? WHERE id = ?', ('已确认', price_id))
        conn.commit()
        return True
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def get_record(id_str: str):
    """Get a single record by ID."""
    conn = open_db()
    try:
        price_id = find_price_id(conn, id_str)
        if not price_id:
            return None
        rows = run_query(conn, """
            SELECT cp.*, ci.name, ci.category
            FROM cost_price cp
            JOIN cost_item ci ON cp.item_id = ci.id
            WHERE cp.id = ?
        """, (price_id,))
        return rows[0] if rows else None
    finally:
        conn.close()


def get_stats() -> dict:
    """Return database statistics."""
    conn = open_db()
    try:
        total = run_query(conn, 'SELECT COUNT(*) AS c FROM cost_price')[0]['c']
        items = run_query(conn, 'SELECT COUNT(*) AS c FROM cost_item')[0]['c']
        by_cat = run_query(conn, """
            SELECT ci.category, COUNT(*) AS c
            FROM cost_price cp JOIN cost_item ci ON cp.item_id = ci.id
            GROUP BY ci.category ORDER BY c DESC
        """)
        by_status = run_query(conn,
            'SELECT status, COUNT(*) AS c FROM cost_price GROUP BY status')
        return {
            'total': total,
            'items': items,
            'by_category': by_cat,
            'by_status': by_status,
        }
    finally:
        conn.close()


def generate_dashboard() -> str:
    """Generate 成本查询.md dashboard. Returns the file path."""
    conn = open_db()
    try:
        now_str = datetime.now().strftime('%Y/%m/%d %H:%M:%S')

        md = '---\n'
        md += 'doc_type: 成本查询\n'
        md += f'updated: "{now_str}"\n'
        md += '---\n\n'
        md += '# 成本查询\n\n'
        md += f'> 数据来源：SQLite（V3.2 关系型）| 最后更新：{now_str}\n'
        md += '> 运行 `python cost_db.py dashboard` 刷新\n\n---\n\n'

        # 1. 单价查询
        md += '## 单价查询\n\n<!-- AUTO-GENERATED: lookup -->\n'
        all_items = run_query(conn, """
            SELECT ci.category, ci.name, cp.unit, COUNT(*) AS samples,
                   ROUND(AVG(cp.price),2) AS avg_price, ROUND(MIN(cp.price),2) AS min_price,
                   ROUND(MAX(cp.price),2) AS max_price, MAX(cp.date) AS latest_date
            FROM cost_price cp JOIN cost_item ci ON cp.item_id = ci.id
            GROUP BY ci.category, ci.name, cp.unit
            ORDER BY latest_date DESC, ci.category, ci.name
        """)
        md += '| 最新日期 | 大类 | 名称 | 单位 | 样本 | 均价 | 最低 | 最高 |\n'
        md += '|----------|------|------|------|------|------|------|------|\n'
        for r in all_items:
            md += f"| {r['latest_date']} | {r['category']} | {r['name']} | {r['unit']} | {r['samples']} | {fmt(r['avg_price'])} | {fmt(r['min_price'])} | {fmt(r['max_price'])} |\n"

        # 2. 最近入库
        md += '\n---\n\n## 最近入库记录\n\n<!-- AUTO-GENERATED: recent -->\n'
        recent = run_query(conn, """
            SELECT cp.date, ci.category, ci.name, cp.spec, cp.price, cp.unit,
                   cp.tax_method, cp.location, cp.project_name, cp.source_person,
                   cp.status, cp.conversion_source
            FROM cost_price cp JOIN cost_item ci ON cp.item_id = ci.id
            ORDER BY cp.date DESC, cp.id DESC LIMIT 50
        """)
        md += '| 日期 | 大类 | 名称 | 规格 | 单价 | 单位 | 计税 | 地区 | 项目 | 报价人 | 状态 |\n'
        md += '|------|------|------|------|------|------|------|------|------|--------|------|\n'
        for r in recent:
            price_note = f" {fmt(r['price'])}(←{r['conversion_source']})" if r['conversion_source'] else fmt(r['price'])
            md += f"| {r['date']} | {r['category']} | {r['name']} | {r['spec'] or '-'} | {price_note} | {r['unit']} | {r['tax_method'] or '不详'} | {r['location'] or '-'} | {r['project_name'] or '-'} | {r['source_person'] or '-'} | {r['status'] or '-'} |\n"

        # 3. 三价对比
        md += '\n---\n\n## 三价对比\n\n<!-- AUTO-GENERATED: three-price -->\n'
        tp_raw = run_query(conn, """
            SELECT ci.name, cp.unit, cp.location, cp.price_type, ROUND(AVG(cp.price),2) AS avg_price
            FROM cost_price cp JOIN cost_item ci ON cp.item_id = ci.id
            WHERE cp.price_type IN ('信息价','合同价','电话询价','微信询价','现场询价','询价单','定额')
            GROUP BY ci.name, cp.unit, cp.location, cp.price_type
        """)
        tp_map = {}
        for r in tp_raw:
            key = f"{r['name']}|{r['unit']}|{r['location'] or ''}"
            if key not in tp_map:
                tp_map[key] = {'name': r['name'], 'unit': r['unit'], 'location': r['location'] or '-', 'prices': {}}
            cat = get_price_category(r['price_type'])
            tp_map[key]['prices'][cat] = r['avg_price']
        tp_rows = [v for v in tp_map.values() if len(v['prices']) >= 2]
        if tp_rows:
            md += '| 名称 | 单位 | 地区 | 信息价 | 市场价 | 合同价 | 偏差 |\n'
            md += '|------|------|------|--------|--------|--------|------|\n'
            for r in tp_rows:
                info = r['prices'].get('信息价')
                market = r['prices'].get('市场价')
                contract = r['prices'].get('合同价')
                deviation = '-'
                if info and market:
                    pct = (market - info) / info * 100
                    deviation = f"{'+' if pct > 0 else ''}{pct:.1f}%"
                    if abs(pct) > 15:
                        deviation += ' ⚠️'
                md += f"| {r['name']} | {r['unit']} | {r['location']} | {fmt(info) if info else '-'} | {fmt(market) if market else '-'} | {fmt(contract) if contract else '-'} | {deviation} |\n"
        else:
            md += '> 暂无足够数据进行三价对比\n'

        # 4. 综合报价拆分
        md += '\n---\n\n## 综合报价工料机拆分\n\n<!-- AUTO-GENERATED: breakdown -->\n'
        comp_raw = run_query(conn, """
            SELECT ci.name, cp.unit, cp.id AS price_id, cp.price
            FROM cost_price cp JOIN cost_item ci ON cp.item_id = ci.id
            WHERE cp.is_composite = 1
            ORDER BY ci.name
        """)
        breakdown_map = {}
        for r in comp_raw:
            key = f"{r['name']}|{r['unit']}"
            if key not in breakdown_map:
                breakdown_map[key] = {'name': r['name'], 'unit': r['unit'], 'samples': 0, 'total': 0, 'labor': 0, 'material': 0, 'equipment': 0}
            b = breakdown_map[key]
            b['samples'] += 1
            b['total'] += r['price']
            comps = run_query(conn,
                'SELECT component_type, price FROM cost_component WHERE price_id = ?',
                (r['price_id'],))
            for c in comps:
                if c['component_type'] == '人工':
                    b['labor'] += c['price']
                if c['component_type'] == '材料':
                    b['material'] += c['price']
                if c['component_type'] == '机械':
                    b['equipment'] += c['price']
        breakdown_rows = list(breakdown_map.values())
        if breakdown_rows:
            md += '| 名称 | 单位 | 样本 | 综合 | 人工 | 材料 | 机械 | 人工% | 材料% | 机械% |\n'
            md += '|------|------|------|------|------|------|------|-------|-------|-------|\n'
            for r in breakdown_rows:
                avg_total = r['total'] / r['samples']
                avg_l = r['labor'] / r['samples']
                avg_m = r['material'] / r['samples']
                avg_e = r['equipment'] / r['samples']
                md += f"| {r['name']} | {r['unit']} | {r['samples']} | {fmt(avg_total)} | {fmt(avg_l) if avg_l else '-'} | {fmt(avg_m) if avg_m else '-'} | {fmt(avg_e) if avg_e else '-'} | {f'{avg_l/avg_total*100:.1f}%' if avg_l else '-'} | {f'{avg_m/avg_total*100:.1f}%' if avg_m else '-'} | {f'{avg_e/avg_total*100:.1f}%' if avg_e else '-'} |\n"
        else:
            md += '> 暂无综合报价的工料机拆分数据\n'

        # 5. 按大类价格汇总
        md += '\n---\n\n## 按大类价格汇总\n\n<!-- AUTO-GENERATED: summary -->\n'
        cat_order = ['人工费', '材料费', '机械费', '综合']
        for cat in cat_order:
            rows = run_query(conn, """
                SELECT ci.name, cp.unit, COUNT(*) AS samples, ROUND(AVG(cp.price),2) AS avg_price,
                       ROUND(MIN(cp.price),2) AS min_price, ROUND(MAX(cp.price),2) AS max_price,
                       ROUND(MAX(cp.price)-MIN(cp.price),2) AS spread
                FROM cost_price cp JOIN cost_item ci ON cp.item_id = ci.id
                WHERE ci.category = ? AND cp.status = '已确认'
                GROUP BY ci.name, cp.unit ORDER BY avg_price DESC
            """, (cat,))
            if not rows:
                continue
            md += f'### {cat}\n\n'
            md += '| 名称 | 单位 | 样本 | 均价 | 最低 | 最高 | 价差 |\n'
            md += '|------|------|------|------|------|------|------|\n'
            for r in rows:
                md += f"| {r['name']} | {r['unit']} | {r['samples']} | {fmt(r['avg_price'])} | {fmt(r['min_price'])} | {fmt(r['max_price'])} | {fmt(r['spread'])} |\n"
            md += '\n'

        # 6. 价格趋势
        md += '---\n\n## 价格趋势\n\n<!-- AUTO-GENERATED: trends -->\n'
        trend_raw = run_query(conn, """
            SELECT ci.name, cp.unit, cp.location, cp.date, cp.price
            FROM cost_price cp JOIN cost_item ci ON cp.item_id = ci.id
            WHERE cp.status = '已确认'
            ORDER BY ci.name, cp.unit, cp.location, cp.date ASC
        """)
        trend_map = {}
        for r in trend_raw:
            key = f"{r['name']}|{r['unit']}|{r['location'] or ''}"
            if key not in trend_map:
                trend_map[key] = {'name': r['name'], 'unit': r['unit'], 'location': r['location'] or '-', 'points': []}
            trend_map[key]['points'].append({'date': r['date'], 'price': r['price']})
        trend_items = [t for t in trend_map.values() if len(t['points']) >= 2]
        if trend_items:
            md += '| 名称 | 单位 | 地区 | 数据点 | 最早价 | 最新价 | 变化 | 趋势 | 价格链 |\n'
            md += '|------|------|------|--------|--------|--------|------|------|--------|\n'
            for t in trend_items:
                first = t['points'][0]
                last = t['points'][-1]
                pct = (last['price'] - first['price']) / first['price'] * 100
                arrow = '↑' if pct > 3 else ('↓' if pct < -3 else '→')
                chain = ' → '.join(fmt(p['price']) for p in t['points'])
                md += f"| {t['name']} | {t['unit']} | {t['location']} | {len(t['points'])} | {fmt(first['price'])} | {fmt(last['price'])} | {'+' if pct > 0 else ''}{pct:.1f}% | {arrow} | {chain} |\n"
        else:
            md += '> 暂无足够数据展示价格趋势\n'

        # 7. 按项目分组
        md += '\n---\n\n## 按项目分组\n\n<!-- AUTO-GENERATED: projects -->\n'
        projects = run_query(conn, """
            SELECT cp.project_name, COUNT(*) AS cnt, COUNT(DISTINCT ci.name) AS items,
                   GROUP_CONCAT(DISTINCT ci.category) AS categories
            FROM cost_price cp JOIN cost_item ci ON cp.item_id = ci.id
            GROUP BY cp.project_name ORDER BY cnt DESC
        """)
        md += '| 项目 | 记录数 | 涉及工项 | 费用大类 |\n'
        md += '|------|--------|----------|----------|\n'
        for r in projects:
            md += f"| {r['project_name'] or '未分类'} | {r['cnt']} | {r['items']} | {r['categories'] or '-'} |\n"

        md += '\n> 查询：`python cost_db.py query "SELECT ..."`\n'

        export_json(conn)
        with open(DASHBOARD_PATH, 'w', encoding='utf-8') as f:
            f.write(md)
        print(f'看板已更新：{DASHBOARD_PATH}')
        return DASHBOARD_PATH
    finally:
        conn.close()


# ── Tax Conversion ──

def cmd_convert_tax(id_str, args):
    params = parse_args(args)
    rate = float(params.get('rate', params.get('税率', '9')))
    conn = open_db()
    try:
        price_id = find_price_id(conn, id_str)
        if not price_id:
            raise ValueError(f'未找到记录 {id_str}')

        rows = run_query(conn, """
            SELECT cp.*, ci.name FROM cost_price cp
            JOIN cost_item ci ON cp.item_id = ci.id WHERE cp.id = ?
        """, (price_id,))
        if not rows:
            print('记录不存在')
            return

        r = rows[0]
        tax_method = r['tax_method'] or '不详'
        if tax_method == '含税':
            pre_tax = r['price'] / (1 + rate / 100)
            print(f'#{price_id} {r["name"]} {r["price"]} {r["unit"]} (含税)')
            print(f'  → 税前价: {pre_tax:.2f} {r["unit"]} (税率{rate:.0f}%)')
        elif tax_method == '税前':
            with_tax = r['price'] * (1 + rate / 100)
            print(f'#{price_id} {r["name"]} {r["price"]} {r["unit"]} (税前)')
            print(f'  → 含税价: {with_tax:.2f} {r["unit"]} (税率{rate:.0f}%)')
        else:
            with_tax = r['price'] * (1 + rate / 100)
            pre_tax = r['price'] / (1 + rate / 100)
            print(f'#{price_id} {r["name"]} {r["price"]} {r["unit"]} (计税方式: 不详)')
            print(f'  假设含税 → 税前: {pre_tax:.2f}')
            print(f'  假设税前 → 含税: {with_tax:.2f}')
    finally:
        conn.close()


# ── Unit Conversion ──

def load_conversions(conn):
    """Load conversion formulas from DB table."""
    rows = run_query(conn,
        'SELECT name, from_unit, to_unit, formula, note FROM conversion_formula ORDER BY id')
    return [{'name': r['name'], 'fromUnit': r['from_unit'],
             'toUnit': r['to_unit'], 'formula': r['formula'],
             'note': r['note']} for r in rows]


def cmd_convert(id_str, args):
    params = parse_args(args)
    conn = open_db()
    try:
        price_id = find_price_id(conn, id_str)
        if not price_id:
            raise ValueError(f'未找到记录 {id_str}')

        rows = run_query(conn, """
            SELECT cp.*, ci.name, ci.category FROM cost_price cp
            JOIN cost_item ci ON cp.item_id = ci.id WHERE cp.id = ?
        """, (price_id,))
        if not rows:
            print('记录不存在')
            return

        source = rows[0]
        conversions = load_conversions(conn)
        if not conversions:
            raise FileNotFoundError('未找到换算公式（conversion_formula 表为空）')

        matches = [c for c in conversions if c['fromUnit'] == source['unit']]
        if not matches:
            print(f'未找到适用于"{source["unit"]}"的换算公式')
            for c in conversions:
                print(f"  {c['name']}: {c['fromUnit']} → {c['toUnit']}")
            return

        formula_name = params.get('formula', params.get('公式'))
        target = None
        if formula_name:
            target = next((c for c in matches if c['name'] == formula_name), None)
            if not target:
                raise ValueError(f'未找到公式"{formula_name}"。匹配的公式：\n' +
                    '\n'.join(f"  {c['name']}: {c['formula']}" for c in matches))
        else:
            if len(matches) == 1:
                target = matches[0]
            else:
                print(f'找到 {len(matches)} 个换算公式，请指定 --formula：')
                for c in matches:
                    print(f"  {c['name']}: {c['fromUnit']} → {c['toUnit']} ({c['formula']})")
                return

        price_str = str(source['price'])
        formula_text = target['formula'].replace('单价', price_str)
        converted_price = _safe_eval_formula(formula_text)

        to_unit = normalize_unit(target['toUnit'], conn)
        item_id = find_or_create_item(conn, source['name'], source['category'], to_unit)

        conn.execute("""
            INSERT INTO cost_price
            (item_id, price, unit, date, tax_method, price_type, source_person,
             location, project_name, spec, status, input_device, source_file,
             is_composite, conversion_source, conversion_formula, remark)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            item_id, round(converted_price * 100) / 100, to_unit, source['date'],
            source['tax_method'], source['price_type'], source['source_person'],
            source['location'], source['project_name'], source['spec'], source['status'],
            source['input_device'], source['source_file'], source['is_composite'],
            str(price_id), target['formula'], f'换算自 #{price_id}: {target["formula"]}',
        ))

        new_id = run_query(conn, 'SELECT last_insert_rowid() AS id')[0]['id']
        export_json(conn)
        conn.commit()

        print(f'换算完成：{source["price"]} {source["unit"]} → {converted_price:.2f} {to_unit}')
        print(f'新记录 #{new_id} (换算自 #{price_id}, 公式: {target["formula"]})')
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


# ── Project Report ──

def cmd_project(name):
    conn = open_db()
    try:
        rows = run_query(conn, """
            SELECT cp.id, cp.date, ci.category, ci.name, cp.spec, cp.price, cp.unit,
                   cp.price_type, cp.source_person, cp.remark
            FROM cost_price cp JOIN cost_item ci ON cp.item_id = ci.id
            WHERE cp.project_name LIKE ?
            ORDER BY cp.date DESC, cp.id DESC
        """, (f'%{name}%',))
        summary = run_query(conn, """
            SELECT ci.category, COUNT(*) AS cnt, ROUND(SUM(cp.price),2) AS total
            FROM cost_price cp JOIN cost_item ci ON cp.item_id = ci.id
            WHERE cp.project_name LIKE ?
            GROUP BY ci.category
        """, (f'%{name}%',))

        if not rows:
            print(f'未找到项目含"{name}"的成本数据')
            return

        md = f'# {name} — 项目成本明细\n\n'
        md += f'> 生成时间：{datetime.now().strftime("%Y/%m/%d %H:%M:%S")} | `python cost_db.py project "{name}"`\n\n'
        md += '| 日期 | 大类 | 名称 | 规格 | 单价 | 单位 | 询价方式 | 报价人 | 备注 |\n'
        md += '|------|------|------|------|------|------|----------|--------|------|\n'
        for r in rows:
            md += f"| {r['date']} | {r['category']} | {r['name']} | {r['spec'] or '-'} | {fmt(r['price'])} | {r['unit']} | {r['price_type'] or '-'} | {r['source_person'] or '-'} | {r['remark'] or '-'} |\n"
        md += '\n### 成本汇总\n\n| 大类 | 记录数 | 合计 |\n|------|--------|------|\n'
        for r in summary:
            md += f"| {r['category']} | {r['cnt']} | {fmt(r['total'])} |\n"

        project_dir = os.environ.get('COST_PROJECT_DIR', '')
        if project_dir and os.path.exists(project_dir):
            safe_name = re.sub(r'[\\/:*?"<>|]', '_', name)
            project_file = None
            for f in os.listdir(project_dir):
                if f.endswith('.md') and (safe_name in f or name in f):
                    project_file = os.path.join(project_dir, f)
                    break

            if project_file:
                with open(project_file, 'r', encoding='utf-8') as f:
                    content = f.read()
                marker = '<!-- AUTO-GENERATED: cost_detail -->'
                end_marker = '<!-- /AUTO-GENERATED: cost_detail -->'
                section = f'\n{marker}\n{md}\n{end_marker}\n'
                si = content.find(marker)
                if si != -1:
                    ei = content.find(end_marker, si)
                    if ei != -1:
                        content = content[:si] + section + content[ei + len(end_marker):]
                    else:
                        content += section
                else:
                    content += '\n\n' + section
                with open(project_file, 'w', encoding='utf-8') as f:
                    f.write(content)
                print(f'项目成本已更新：{project_file}')
            else:
                print(md)
        else:
            print(md)
    finally:
        conn.close()


# ── Items Management ──

def cmd_items(args):
    subcmd = args[0] if args else 'list'
    conn = open_db()
    try:
        if subcmd == 'list':
            rows = run_query(conn, """
                SELECT ci.id, ci.name, ci.category, ci.aliases,
                  (SELECT COUNT(*) FROM cost_price WHERE item_id = ci.id) AS samples
                FROM cost_item ci ORDER BY ci.category, ci.name
            """)
            print(f'共 {len(rows)} 个成本项：\n')
            last_cat = ''
            for r in rows:
                if r['category'] != last_cat:
                    last_cat = r['category']
                    print(f'[{last_cat}]')
                alias = f" (别名: {r['aliases']})" if r['aliases'] else ''
                print(f"  #{r['id']} {r['name']} — {r['samples']}条{alias}")
        elif subcmd == 'add':
            params = parse_args(args[1:])
            name = params.get('名称') or params.get('name')
            cat = params.get('大类') or params.get('category')
            if not name or not cat:
                print('用法：items add --名称 X --大类 X [--aliases X]')
                sys.exit(1)
            aliases = params.get('aliases') or params.get('别名') or ''
            try:
                conn.execute('INSERT INTO cost_item (name, category, aliases) VALUES (?, ?, ?)',
                             (name, cat, aliases))
                conn.commit()
                print(f'已添加：{name} ({cat})')
            except Exception as e:
                if 'UNIQUE' in str(e):
                    print(f'已存在：{name} ({cat})')
                else:
                    raise
        else:
            print('子命令：list | add')
    finally:
        conn.close()


# ── Units Management ──

def cmd_units(args):
    subcmd = args[0] if args else 'list'
    conn = open_db()
    try:
        if subcmd == 'list':
            rows = run_query(conn, """
                SELECT standard_form, GROUP_CONCAT(unit) AS aliases, precision_rule
                FROM unit_standard GROUP BY standard_form
                ORDER BY precision_rule, standard_form
            """)
            print(f'共 {len(rows)} 个标准单位：\n')
            for r in rows:
                alias_list = [a for a in r['aliases'].split(',') if a != r['standard_form']]
                alias_str = f" (别名: {', '.join(alias_list)})" if alias_list else ''
                print(f"  {r['standard_form']}  [{r['precision_rule']}]{alias_str}")
        elif subcmd == 'add':
            params = parse_args(args[1:])
            unit = params.get('unit')
            standard = params.get('standard')
            precision = params.get('precision')
            if not unit or not standard or not precision:
                print('用法：units add --unit <单位> --standard <标准> --precision <精度>')
                sys.exit(1)
            conn.execute(
                'INSERT OR REPLACE INTO unit_standard (unit, standard_form, precision_rule) VALUES (?, ?, ?)',
                (unit, standard, precision))
            conn.commit()
            global _unit_map
            _unit_map = None
            print(f'已添加：{unit} → {standard}')
        elif subcmd == 'normalize':
            unit = args[1] if len(args) > 1 else None
            if not unit:
                print('用法：units normalize <单位>')
                sys.exit(1)
            print(normalize_unit(unit, conn))
        else:
            print('子命令：list | add | normalize')
    finally:
        conn.close()


# ── CLI Helpers ──

def parse_args(args):
    params = {}
    i = 0
    while i < len(args):
        if args[i].startswith('--'):
            key = args[i][2:]
            if i + 1 < len(args) and not args[i + 1].startswith('--'):
                params[key] = args[i + 1]
                i += 2
            else:
                params[key] = True
                i += 1
        else:
            i += 1
    return params


# ── Pending Excel (待审核中转) ──

# Excel 列定义：列标题 → (参数中文名, 是否必填)
PENDING_COLUMNS = [
    ('批次号', False),
    ('序号', False),
    ('日期', True),
    ('大类', True),
    ('名称', True),
    ('规格', False),
    ('单价', True),
    ('单位', True),
    ('计税方式', False),
    ('询价方式', False),
    ('报价人', False),
    ('地区', False),
    ('项目', False),
    ('录入设备', False),
    ('原始文件', False),
    ('人工费', False),
    ('材料费', False),
    ('机械费', False),
    ('备注', False),
    ('校验结果', False),
    ('审核状态', False),
    ('入库结果', False),
]


def _get_pending_path(date_str=None):
    """Get the pending Excel file path for a given date."""
    if not date_str:
        date_str = datetime.now().strftime('%Y-%m-%d')
    return os.path.join(PENDING_DIR, f'待审核_{date_str}.xlsx')


def _init_pending_wb(ws):
    """Initialize a pending worksheet with headers and formatting."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill('solid', fgColor='FFF2CC')
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))

    for col_idx, (title, _) in enumerate(PENDING_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Column widths
    col_widths = {
        '批次号': 20, '序号': 6, '日期': 12, '大类': 10, '名称': 12,
        '规格': 12, '单价': 10, '单位': 10, '计税方式': 10, '询价方式': 12,
        '报价人': 10, '地区': 10, '项目': 16, '录入设备': 10, '原始文件': 16,
        '人工费': 10, '材料费': 10, '机械费': 10, '备注': 20,
        '校验结果': 30, '审核状态': 10, '入库结果': 16,
    }
    for col_idx, (title, _) in enumerate(PENDING_COLUMNS, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = \
            col_widths.get(title, 12)

    # Data validation for 审核状态 column (U)
    status_col = 21  # 审核状态 is the 21st column
    status_letter = ws.cell(row=1, column=status_col).column_letter
    dv = DataValidation(type='list', formula1='"待审核,已审核,已拒绝"', allow_blank=False)
    dv.error = '请选择：待审核、已审核、已拒绝'
    dv.errorTitle = '无效输入'
    ws.add_data_validation(dv)
    dv.add(f'{status_letter}2:{status_letter}1000')

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = 'C2'


def pending_record(params: dict) -> str:
    """Write pending data to Excel for review. Returns the file path."""
    import openpyxl
    from openpyxl.styles import PatternFill

    name = params.get('名称') or params.get('name')
    category = params.get('大类') or params.get('category')
    price = float(params.get('单价') or params.get('price') or 0)
    unit = params.get('单位') or params.get('unit')
    date = params.get('日期') or params.get('date')

    if not name or not category or not price or not unit or not date:
        raise ValueError('缺少必填字段（日期/大类/名称/单价/单位）')

    # Validate and normalize via DB (read-only)
    conn = open_db()
    try:
        normed_unit = normalize_unit(unit, conn)

        # Run validations (read-only) — gracefully handle missing tables
        validation_warnings = []
        try:
            validation_warnings.extend(validate_price(conn, name, normed_unit, price, category))
        except Exception:
            pass  # validation_rule table may not exist in older DBs
        # For duplicate/trend checks, we need item_id (may not exist yet)
        existing_items = run_query(conn,
            'SELECT id FROM cost_item WHERE name = ? AND category = ?',
            (name, category))
        if existing_items:
            item_id = existing_items[0]['id']
            source_person = params.get('报价人') or params.get('source_person') or ''
            location = params.get('地区') or params.get('location') or ''
            try:
                validation_warnings.extend(check_duplicate(conn, item_id, date, price, source_person))
            except Exception:
                pass
            try:
                validation_warnings.extend(check_trend(conn, name, normed_unit, location, price))
            except Exception:
                pass

        # Check breakdown
        labor = params.get('人工费')
        material = params.get('材料费')
        equipment = params.get('机械费')
        labor = float(labor) if labor else None
        material = float(material) if material else None
        equipment = float(equipment) if equipment else None
        if labor is not None or material is not None or equipment is not None:
            s = (labor or 0) + (material or 0) + (equipment or 0)
            if price > 0 and abs(s - price) / price > 0.05:
                pct = (s - price) / price * 100
                validation_warnings.append(
                    f'工料机合计{s}与单价偏差{pct:.1f}%')

        # Check if name in validation rules
        try:
            rule_check = run_query(conn,
                'SELECT id FROM validation_rule WHERE name = ? LIMIT 1', (name,))
            if not rule_check:
                validation_warnings.append('待补充词条')
        except Exception:
            pass  # validation_rule table may not exist
    finally:
        conn.close()

    validation_str = '; '.join(validation_warnings) if validation_warnings else ''

    # Ensure pending dir exists
    os.makedirs(PENDING_DIR, exist_ok=True)

    batch_no = 'P-' + datetime.now().strftime('%Y%m%d-%H%M%S')
    excel_path = _get_pending_path(date)

    # Write to Excel with retry for concurrent access
    max_retries = 3
    for attempt in range(max_retries):
        try:
            if os.path.exists(excel_path):
                wb = openpyxl.load_workbook(excel_path)
                ws = wb.active
                row_idx = ws.max_row + 1
                # Determine sequence number
                seq = 1
                for r in range(2, ws.max_row + 1):
                    if ws.cell(row=r, column=1).value and \
                       ws.cell(row=r, column=1).value.startswith(batch_no[:11]):
                        seq += 1
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = '待审核'
                _init_pending_wb(ws)
                row_idx = 2
                seq = 1

            # Column values in order
            values = [
                batch_no,                                    # A 批次号
                seq,                                         # B 序号
                date,                                        # C 日期
                category,                                    # D 大类
                name,                                        # E 名称
                params.get('规格') or params.get('spec') or '',  # F 规格
                price,                                       # G 单价
                normed_unit,                                 # H 单位
                params.get('计税方式') or params.get('tax_method') or '不详',  # I
                params.get('询价方式') or params.get('price_type') or '',     # J
                params.get('报价人') or params.get('source_person') or '',   # K
                params.get('地区') or params.get('location') or '',         # L
                params.get('项目') or params.get('project_name') or '',     # M
                params.get('录入设备') or params.get('input_device') or 'QQ',  # N
                params.get('原始文件') or params.get('source_file') or '',   # O
                labor if labor is not None else '',           # P 人工费
                material if material is not None else '',     # Q 材料费
                equipment if equipment is not None else '',   # R 机械费
                params.get('备注') or params.get('remark') or '',  # S 备注
                validation_str,                              # T 校验结果
                '待审核',                                     # U 审核状态
                '',                                          # V 入库结果
            ]

            warn_fill = PatternFill('solid', fgColor='FCE4EC')
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                if validation_warnings:
                    cell.fill = warn_fill

            wb.save(excel_path)
            break
        except (PermissionError, OSError):
            if attempt < max_retries - 1:
                time.sleep(0.5)
            else:
                raise

    result_msg = f'已写入待审核文件：{excel_path}'
    if validation_warnings:
        result_msg += f'\n  校验警告：{validation_str}'
    print(result_msg)
    return excel_path


def commit_pending(excel_path=None) -> dict:
    """Read reviewed Excel, commit approved rows to SQLite.
    Returns stats dict.
    """
    import openpyxl

    if not excel_path:
        # Find latest pending file
        if not os.path.exists(PENDING_DIR):
            raise FileNotFoundError('待审核目录不存在，请先执行 pending 命令')
        files = sorted(
            [f for f in os.listdir(PENDING_DIR) if f.startswith('待审核_') and f.endswith('.xlsx')],
            reverse=True)
        if not files:
            raise FileNotFoundError('未找到待审核文件')
        excel_path = os.path.join(PENDING_DIR, files[0])

    if not os.path.exists(excel_path):
        raise FileNotFoundError(f'文件不存在：{excel_path}')

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # Read header row to build column index
    headers = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if val:
            headers[val] = col_idx

    required_headers = ['日期', '大类', '名称', '单价', '单位', '审核状态', '入库结果']
    for h in required_headers:
        if h not in headers:
            raise ValueError(f'Excel 缺少必要列：{h}')

    stats = {'total': 0, 'committed': 0, 'skipped': 0, 'errors': [], 'details': []}

    for row_idx in range(2, ws.max_row + 1):
        # Skip rows already processed
        result_col = headers.get('入库结果')
        if result_col:
            existing_result = ws.cell(row=row_idx, column=result_col).value
            if existing_result and str(existing_result).strip():
                stats['skipped'] += 1
                continue

        status_val = ws.cell(row=row_idx, column=headers['审核状态']).value
        if status_val != '已审核':
            stats['skipped'] += 1
            continue

        stats['total'] += 1

        # Build params dict
        params = {}
        col_map = {
            '日期': '日期', '大类': '大类', '名称': '名称', '单价': '单价',
            '单位': '单位', '计税方式': '计税方式', '询价方式': '询价方式',
            '报价人': '报价人', '地区': '地区', '项目': '项目', '规格': '规格',
            '录入设备': '录入设备', '原始文件': '原始文件', '备注': '备注',
        }
        for excel_col, param_key in col_map.items():
            if excel_col in headers:
                val = ws.cell(row=row_idx, column=headers[excel_col]).value
                if val is not None and str(val).strip():
                    params[param_key] = str(val).strip() if not isinstance(val, (int, float)) else val

        # Component breakdown
        for comp_key in ['人工费', '材料费', '机械费']:
            if comp_key in headers:
                val = ws.cell(row=row_idx, column=headers[comp_key]).value
                if val is not None and str(val).strip():
                    params[comp_key] = float(val)

        # Append validation result from Excel to remark
        if '校验结果' in headers:
            val = ws.cell(row=row_idx, column=headers['校验结果']).value
            if val and str(val).strip():
                remark = params.get('备注', '')
                params['备注'] = f'{remark} [Excel校验: {val}]'.strip()

        try:
            new_id = insert_record(params)
            stats['committed'] += 1
            stats['details'].append({'row': row_idx, 'id': new_id, 'status': 'ok'})
            # Update Excel
            if result_col:
                ws.cell(row=row_idx, column=result_col, value=f'已入库 #{new_id}')
        except Exception as e:
            stats['errors'].append({'row': row_idx, 'error': str(e)})
            stats['details'].append({'row': row_idx, 'status': 'error', 'error': str(e)})
            if result_col:
                ws.cell(row=row_idx, column=result_col, value=f'失败: {e}')

    wb.save(excel_path)

    print(f'审核入库完成：')
    print(f'  已入库：{stats["committed"]} 条')
    print(f'  跳过：{stats["skipped"]} 条')
    if stats['errors']:
        print(f'  失败：{len(stats["errors"])} 条')
        for e in stats['errors']:
            print(f'    第{e["row"]}行：{e["error"]}')
    return stats


def cmd_pending_list():
    """List all pending Excel files with stats."""
    import openpyxl

    if not os.path.exists(PENDING_DIR):
        print('待审核目录不存在')
        return

    files = sorted(
        [f for f in os.listdir(PENDING_DIR) if f.startswith('待审核_') and f.endswith('.xlsx')],
        reverse=True)

    if not files:
        print('暂无待审核文件')
        return

    print(f'共 {len(files)} 个待审核文件：\n')
    for f in files:
        fpath = os.path.join(PENDING_DIR, f)
        try:
            wb = openpyxl.load_workbook(fpath, read_only=True)
            ws = wb.active
            total_rows = ws.max_row - 1  # exclude header
            pending = 0
            approved = 0
            rejected = 0
            committed = 0
            # Read headers
            headers = {}
            for col_idx in range(1, ws.max_column + 1):
                val = ws.cell(row=1, column=col_idx).value
                if val:
                    headers[val] = col_idx
            for row_idx in range(2, ws.max_row + 1):
                status = ws.cell(row=row_idx, column=headers.get('审核状态', 21)).value or ''
                result = ws.cell(row=row_idx, column=headers.get('入库结果', 22)).value or ''
                if str(result).strip():
                    committed += 1
                elif status == '已审核':
                    approved += 1
                elif status == '已拒绝':
                    rejected += 1
                else:
                    pending += 1
            wb.close()
            print(f'  {f}')
            print(f'    总计 {total_rows} 条 | 待审核 {pending} | 已审核 {approved} | 已拒绝 {rejected} | 已入库 {committed}')
        except Exception as e:
            print(f'  {f} (读取失败: {e})')


# ── Main CLI ──

def main():
    argv = sys.argv[1:]
    if not argv:
        print('成本数据库工具 V3.2\n')
        print('用法：python cost_db.py <command> [args]\n')
        print('命令：')
        print('  pending --日期 ...   写入待审核 Excel（QQ Bot 入库推荐）')
        print('  commit [--file X]    将已审核数据导入 SQLite')
        print('  pending-list         列出待审核文件')
        print('  insert --日期 ...    直接插入记录（手动/调试用）')
        print('  update <id> <f> <v>  更新记录')
        print('  delete <id>          删除记录（级联删除）')
        print('  query "<sql>"        SQL 查询')
        print('  dashboard            生成看板')
        print('  project "<name>"     项目成本报表')
        print('  stats                统计')
        print('  convert-tax <id>     含税/税前换算')
        print('  convert <id>         单位换算')
        print('  items [list|add]     成本项管理')
        print('  units [list|add|...] 单位管理')
        return

    cmd = argv[0]
    rest = argv[1:]

    try:
        if cmd == 'pending':
            params = parse_args(rest)
            pending_record(params)
        elif cmd == 'commit':
            p = parse_args(rest)
            if p.get('list'):
                cmd_pending_list()
            elif p.get('all'):
                # Commit all pending files
                if os.path.exists(PENDING_DIR):
                    files = sorted(f for f in os.listdir(PENDING_DIR)
                                   if f.startswith('待审核_') and f.endswith('.xlsx'))
                    for f in files:
                        fpath = os.path.join(PENDING_DIR, f)
                        print(f'\n--- 处理：{f} ---')
                        commit_pending(fpath)
                else:
                    print('待审核目录不存在')
            else:
                commit_pending(p.get('file'))
        elif cmd == 'pending-list':
            cmd_pending_list()
        elif cmd == 'insert':
            params = parse_args(rest)
            print('[注意] insert 已废弃，自动转为 pending 写入待审核 Excel')
            pending_record(params)
        elif cmd == 'update':
            if len(rest) < 3:
                print('用法：update <id> <字段> <值>')
                sys.exit(1)
            update_record(rest[0], rest[1], ' '.join(rest[2:]))
        elif cmd == 'delete':
            if not rest:
                print('用法：delete <id>')
                sys.exit(1)
            delete_record(rest[0])
        elif cmd == 'query':
            if not rest:
                print('用法：query "<sql>"')
                sys.exit(1)
            rows = query_prices(' '.join(rest))
            print(json.dumps(rows, ensure_ascii=False, indent=2))
        elif cmd == 'dashboard':
            generate_dashboard()
        elif cmd == 'project':
            if not rest:
                print('用法：project "<name>"')
                sys.exit(1)
            cmd_project(' '.join(rest))
        elif cmd == 'stats':
            s = get_stats()
            print(f"总记录：{s['total']}")
            print(f"成本项：{s['items']}")
            print('\n按大类：')
            for r in s['by_category']:
                print(f"  {r['category']}: {r['c']}")
            print('\n按状态：')
            for r in s['by_status']:
                print(f"  {r['status']}: {r['c']}")
        elif cmd == 'convert-tax':
            if not rest:
                print('用法：convert-tax <id> [--rate 9]')
                sys.exit(1)
            cmd_convert_tax(rest[0], rest[1:])
        elif cmd == 'convert':
            if not rest:
                print('用法：convert <id> [--formula <名>]')
                sys.exit(1)
            cmd_convert(rest[0], rest[1:])
        elif cmd == 'items':
            cmd_items(rest)
        elif cmd == 'units':
            cmd_units(rest)
        else:
            print(f'未知命令：{cmd}')
            sys.exit(1)
    except Exception as e:
        print(f'错误：{e}', file=sys.stderr)
        sys.exit(1)


if __name__ == '__main__':
    main()
